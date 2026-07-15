#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Combinaisons de piliers DORA : sortie QUALITATIVE, l'ORDRE pilote fortement.

Principe : une cascade qui suit une direction causale PLAUSIBLE (amont → aval,
gouvernance → conséquences) est un vrai emballement systémique → plus PROBABLE
ET plus GRAVE. Le même ensemble de piliers pris à l'envers est peu probable et,
s'il survient, bien moins grave (défaillances quasi indépendantes, non compoundées).

=> L'ordre change fortement le verdict (ex. 1→2 « Majeure » vs 2→1 « Faible »).

Résultat exprimé en CATÉGORIES (qualitatif), pas en nombres bruts :
  Proba    : Très rare · Rare · Possible · Probable · Très probable
  Gravité  : Négligeable · Mineure · Modérée · Majeure · Critique
  Criticité: Faible · Modérée · Élevée · Majeure · Extrême   (croisement des deux)
Les scores /10 restent en colonnes d'appoint (grises) juste pour trier.
"""

import os
import sys
from itertools import combinations, permutations
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- jugement et barème : source unique (cascade_model.py) ------------------
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cascade_model import (PILIERS, N,
                           PROBA_LAB, GRAV_LAB, CRIT_LAB,
                           lvl, proba_score, gravite_score, crit_score)

# vert -> rouge (échelle de couleurs, spécifique au classeur)
LEVEL_FILL = ["C6EFCE", "E2EFDA", "FFEB9C", "FCE4D6", "FFC7CE"]


# --- styles ----------------------------------------------------------------
WHITE_BOLD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="4472C4")
GREY = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
K_FILL = {0: "D9D9D9", 1: "E2EFDA", 2: "DDEBF7", 3: "FFF2CC", 4: "FCE4D6", 5: "F8CBAD"}


def fill(c):
    return PatternFill("solid", fgColor=c)


def head(cell):
    cell.font = WHITE_BOLD
    cell.fill = HEAD_FILL
    cell.alignment = CENTER
    cell.border = BORDER


LATIN = ["", "bis", "ter", "quater", "quinquies", "sexies", "septies", "octies",
         "novies", "decies", "undecies", "duodecies", "terdecies", "quaterdecies",
         "quindecies", "sexdecies", "septdecies", "octodecies", "novdecies", "vicies"]


def suffixe(p):
    return LATIN[p - 1] if p - 1 < len(LATIN) else f".{p}"


wb = Workbook()

# ==========================================================================
# Feuille 1 : Combinaisons
# ==========================================================================
ws = wb.active
ws.title = "Combinaisons"
ws["A1"] = "Combinaisons de piliers DORA : verdict qualitatif, l'ORDRE pilote"
ws["A1"].font = Font(bold=True, size=13)
ws["A2"] = ("Piliers : " + "  |  ".join(f"P{j} = {PILIERS[j]}" for j in range(1, N + 1))
            + " : cascade cohérente (amont→aval) = probable ET grave ; à l'envers = rare et peu grave.")
ws["A2"].font = Font(italic=True, color="808080")

HEAD_ROW = 4
headers = ["Réf", "N° combo", "Taille k", "Code binaire\n(P1 P2 P3 P4 P5)",
           "Ordre (chaîne)", "Proba conceptuelle", "Gravité conceptuelle",
           "Criticité", "p/10", "g/10", "crit"]
for idx, h in enumerate(headers, start=1):
    head(ws.cell(row=HEAD_ROW, column=idx, value=h))

records = []
combo_num = 0
for k in range(0, N + 1):
    for subset in combinations(range(1, N + 1), k):
        combo_num += 1
        bincode = " ".join("1" if p in subset else "0" for p in range(1, N + 1))
        perms = [()] if k == 0 else list(permutations(subset))
        for p_idx, order in enumerate(perms, start=1):
            ref = f"{combo_num}{suffixe(p_idx)}"
            ordre = "-" if k == 0 else "→".join(str(x) for x in order)
            ps, gs = proba_score(order), gravite_score(order)
            cs = crit_score(ps, gs)
            records.append((ref, combo_num, k, bincode, ordre, ps, gs, cs))

first = HEAD_ROW + 1
for i, (ref, cnum, k, bincode, ordre, ps, gs, cs) in enumerate(records):
    r = first + i
    # colonnes descriptives : bande de couleur par k
    for col, v in enumerate([ref, cnum, k, bincode, ordre], start=1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.border = BORDER
        cell.fill = fill(K_FILL[k])
        cell.alignment = CENTER
    ws.cell(row=r, column=1).font = Font(bold=True)
    # colonnes qualitatives : label + couleur verte->rouge selon le niveau
    for col, (score, labs) in zip((6, 7, 8), [(ps, PROBA_LAB), (gs, GRAV_LAB), (cs, CRIT_LAB)]):
        li = lvl(score)
        cell = ws.cell(row=r, column=col, value=(labs[li] if li is not None else "-"))
        cell.border = BORDER
        cell.alignment = CENTER
        cell.font = Font(bold=(col == 8))
        cell.fill = fill(LEVEL_FILL[li]) if li is not None else fill("D9D9D9")
    # colonnes d'appoint numériques (grises) pour trier
    for col, v in zip((9, 10, 11), [ps, gs, cs]):
        cell = ws.cell(row=r, column=col, value=v)
        cell.border = BORDER
        cell.alignment = CENTER
        cell.fill = GREY
        cell.font = Font(color="808080")

ws.freeze_panes = f"A{first}"
for i, w in enumerate([12, 9, 7, 15, 14, 16, 17, 13, 6, 6, 6], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ==========================================================================
# Feuille 2 : Justification
# ==========================================================================
js = wb.create_sheet("Justification")
js["A1"] = "Justification : comment l'ordre pilote le verdict (ancré DORA / mémoire)"
js["A1"].font = Font(bold=True, size=13)
notes = [
    "",
    ("PROBA (dépend de l'ordre)", True),
    ("score = MAX(1 ; MIN(10 ; round(10 × ROOT[départ] × Π TRANS[maillon]))). Cascade longue ⇒ plus rare.", False),
    ("ROOT = qui amorce : P1=1.0 (gouvernance, cause racine), P4=0.9 (tiers/cloud), P2=0.6, P3=0.5, P5=0.3.", False),
    ("TRANS = force dirigée « i entraîne j », ASYMÉTRIQUE : P1→X fort, X→P1 faible, P4→P2 fort, P5 faible partout.", False),
    ("", False),
    ("GRAVITÉ (dépend de l'ordre : c'est le changement clé)", True),
    ("étendue = MIN(10 ; max(GBASE) + 0.4 × somme(GBASE des autres piliers du combo)).", False),
    ("gravité = étendue × amplification, avec amplification = 0.5 + 0.8 × cohérence de l'ordre.", False),
    ("cohérence = moyenne des TRANS le long de la chaîne : ordre PLAUSIBLE (amont→aval) ⇒ ×1.3 (vrai emballement) ;", False),
    ("ordre IMPLAUSIBLE ⇒ ×0.5 (défaillances quasi indépendantes, non compoundées).", False),
    ("GBASE (gravité d'un pilier) : P4=7 (systémique, contagion externe cloud rang-3) > P1=6 (fondation, sanctions)", False),
    ("   > P2=4 (notification ACPR) > P3=3 (impréparation latente) > P5=1 (« moins mobilisable », mémoire).", False),
    ("", False),
    ("CRITICITÉ = moyenne géométrique( proba , gravité ) : croise likelihood et impact.", True),
    ("", False),
    ("Effet de l'ordre : exemple {P1,P2} :", True),
    ("  1→2 (gouvernance ⇒ incident) : Proba « Probable », Gravité « Critique », Criticité « Majeure ».", False),
    ("  2→1 (incident ⇒ gouvernance) : Proba « Très rare », Gravité « Modérée », Criticité « Faible ».", False),
    ("(La criticité plafonne à « Majeure » : proba et gravité étant anti-corrélées, aucun scénario n'est à la fois très probable ET très grave.)", False),
]
r = 3
for item in notes:
    txt, bold = item if isinstance(item, tuple) else (item, False)
    js[f"A{r}"] = txt
    js[f"A{r}"].font = Font(bold=True) if bold else Font(color="404040")
    r += 1
js.column_dimensions["A"].width = 115

# ==========================================================================
# Feuille 3 : Classement (325 scénarios notés triés par Criticité décroissante)
# ==========================================================================
cl = wb.create_sheet("Classement", 1)
cl["A1"] = "Scénarios classés par Criticité décroissante (l'ordre compte)"
cl["A1"].font = Font(bold=True, size=13)
cl_head = ["Rang", "Réf", "Ordre", "Code binaire", "Proba conceptuelle",
           "Gravité conceptuelle", "Criticité", "crit"]
for idx, h in enumerate(cl_head, start=1):
    head(cl.cell(row=3, column=idx, value=h))

# tri : criticité, puis gravité, puis proba (décroissant) ; on exclut le « sain »
tri = sorted([rec for rec in records if rec[7] is not None],
             key=lambda x: (x[7], x[6], x[5]), reverse=True)
for rank, (ref, cnum, k, bincode, ordre, ps, gs, cs) in enumerate(tri, start=1):
    r = 3 + rank
    cl.cell(row=r, column=1, value=rank)
    cl.cell(row=r, column=2, value=ref).font = Font(bold=True)
    cl.cell(row=r, column=3, value=ordre)
    cl.cell(row=r, column=4, value=bincode)
    for col, (score, labs) in zip((5, 6, 7), [(ps, PROBA_LAB), (gs, GRAV_LAB), (cs, CRIT_LAB)]):
        li = lvl(score)
        c = cl.cell(row=r, column=col, value=labs[li])
        c.fill = fill(LEVEL_FILL[li])
        c.font = Font(bold=(col == 7))
    cl.cell(row=r, column=8, value=cs).font = Font(color="808080")
    for col in range(1, 9):
        cl.cell(row=r, column=col).border = BORDER
        cl.cell(row=r, column=col).alignment = CENTER
cl.freeze_panes = "A4"
for i, w in enumerate([6, 11, 14, 14, 16, 17, 13, 6], start=1):
    cl.column_dimensions[get_column_letter(i)].width = w

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cascade_piliers_DORA.xlsx")
wb.save(out)
print(f"OK : {len(records)} lignes ; verdict qualitatif piloté par l'ordre -> {out}")
